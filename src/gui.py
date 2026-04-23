#!/usr/bin/env python3
"""
펌프 용량 계산기 GUI 프로그램
tkinter를 사용한 사용자 친화적인 인터페이스
"""

import tkinter as tk
from tkinter import ttk
import sys
import os

# 상대 임포트 대신 절대 경로로 필요한 모듈 임포트
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from calc.pump_calc import calculate_complete_pump_spec


class PumpCalculatorGUI:
    """펌프 용량 계산기 GUI 클래스"""

    def __init__(self, root):
        """GUI 초기화"""
        self.root = root
        self.root.title("하수도 펌프 용량 계산기")
        self.root.geometry("1100x850")
        self.root.resizable(True, True)

        # 변수 초기화
        self.mode_var = tk.StringVar(value='standard')
        self.electrical_var = tk.StringVar(value='3phase')

        # 입력 필드 변수
        self.flow_var = tk.StringVar()
        self.head_var = tk.StringVar()
        self.pumps_var = tk.IntVar(value=1)

        # 압송 모드 변수
        self.pipe_loss_var = tk.StringVar(value='0')
        self.pipe_length_var = tk.StringVar()
        self.pipe_diameter_var = tk.StringVar()
        self.pipe_material_var = tk.StringVar(value='PVC')

        # 옵션 변수
        self.safety_var = tk.StringVar(value='1.5')
        self.efficiency_var = tk.StringVar(value='0.70')
        self.motor_safety_var = tk.StringVar(value='1.15')

        # 메인 레이아웃 생성
        self.create_main_layout()

        # 모드 변경 이벤트 바인딩
        self.mode_var.trace('w', self.on_mode_changed)

    def create_main_layout(self):
        """메인 레이아웃 생성"""
        # 타이틀 영역
        self.create_title()
        
        # 메인 컨테이너 (스크롤 가능)
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # 캔버스와 스크롤바 생성
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        
        # 스크롤 프레임 업데이트 함수
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        scroll_frame.bind("<Configure>", configure_scroll_region)
        
        # UI 생성
        self.create_left_panel(scroll_frame)
        
        # 파일 저장 관련 변수 초기화
        import os
        self.save_path_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop"))
        self.file_format_var = tk.StringVar(value="excel")

    def create_title(self):
        """타이틀 영역 생성"""
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.pack(fill=tk.X)

        title_label = tk.Label(
            title_frame,
            text="하수도 펌프 용량 계산기",
            font=("Helvetica", 20, "bold")
        )
        title_label.pack()

        subtitle_label = tk.Label(
            title_frame,
            text="일반 양정 모드와 압송 관로 모드 지원",
            font=("Helvetica", 13)
        )
        subtitle_label.pack()

    
    def create_file_save_section(self, parent):
        """파일 저장 설정 영역 생성"""
        save_frame = ttk.Frame(parent)
        save_frame.pack(fill=tk.X, padx=10, pady=15)
        
        # 제목 라벨
        title_label = tk.Label(
            save_frame,
            text="파일 저장 설정",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(anchor=tk.W, padx=5, pady=(0, 10))
        
        # 내용 프레임
        content_frame = ttk.Frame(save_frame)
        content_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 저장 경로 선택
        tk.Label(content_frame, text="저장 위치:", font=("Helvetica", 13)).grid(
            row=0, column=0, padx=8, pady=8, sticky=tk.W
        )
        
        # 경로 표시 Entry
        path_entry = tk.Entry(content_frame, textvariable=self.save_path_var, width=40, font=("Helvetica", 12))
        path_entry.grid(row=0, column=1, padx=8, pady=8)
        
        # 폴더 선택 버튼
        browse_button = tk.Button(
            content_frame,
            text="폴더 선택",
            command=self.browse_save_folder,
            bg="#95a5a6",
            fg="white",
            font=("Helvetica", 12),
            padx=15,
            pady=6
        )
        browse_button.grid(row=0, column=2, padx=8, pady=8)
        
        # 파일 형식 선택
        tk.Label(content_frame, text="파일 형식:", font=("Helvetica", 13)).grid(
            row=1, column=0, padx=8, pady=8, sticky=tk.W
        )
        
        # 라디오 버튼 프레임
        format_frame = ttk.Frame(content_frame)
        format_frame.grid(row=1, column=1, columnspan=2, padx=8, pady=8, sticky=tk.W)
        
        # Excel 형식
        excel_radio = tk.Radiobutton(
            format_frame,
            text="Excel 파일 (.xlsx)",
            variable=self.file_format_var,
            value="excel",
            font=("Helvetica", 12)
        )
        excel_radio.pack(side=tk.LEFT, padx=15)
        
        # 텍스트 형식
        text_radio = tk.Radiobutton(
            format_frame,
            text="텍스트 파일 (.txt)",
            variable=self.file_format_var,
            value="text",
            font=("Helvetica", 12)
        )
        text_radio.pack(side=tk.LEFT, padx=15)
        
        return save_frame
    
    def browse_save_folder(self):
        """저장 폴더 선택 다이얼로그"""
        from tkinter import filedialog
        folder_path = filedialog.askdirectory(
            title="결과 파일 저장 폴더 선택",
            initialdir=self.save_path_var.get()
        )
        if folder_path:
            self.save_path_var.set(folder_path)
    
    def save_calculation_result(self, result):
        """계산 결과를 파일로 저장"""
        try:
            import os
            import datetime
            
            # 파일명 생성
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            mode_name = "일반양정" if result["mode"] == "standard" else "압송관로"
            
            if self.file_format_var.get() == "excel":
                filename = f"펌프계산_{mode_name}_{timestamp}.xlsx"
                filepath = os.path.join(self.save_path_var.get(), filename)
                self.save_to_excel(result, filepath)
            else:
                filename = f"펌프계산_{mode_name}_{timestamp}.txt"
                filepath = os.path.join(self.save_path_var.get(), filename)
                self.save_to_text(result, filepath)
            
            # 저장 완료 메시지
            from tkinter import messagebox
            messagebox.showinfo("저장 완료", f"계산 결과가 저장되었습니다:\n{filepath}")
            
            return True
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("저장 오류", f"파일 저장 중 오류가 발생했습니다:\n{str(e)}")
            return False
    
    def save_to_excel(self, result, filepath):
        """Excel 파일로 저장"""
        try:
            # openpyxl이 설치되어 있는지 확인
            try:
                import openpyxl
                from openpyxl import Workbook
            except ImportError:
                # openpyxl이 없으면 설치하도록 안내
                from tkinter import messagebox
                messagebox.showwarning(
                    "라이브러리 필요",
                    "Excel 저장을 위해 openpyxl 라이브러리가 필요합니다.\n\n설치 명령: pip install openpyxl"
                )
                return False
            
            # 새 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "펌프 계산 결과"
            
            # 제목
            ws["A1"] = "하수도 펌프 용량 계산 결과"
            ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
            
            # 기본 정보
            ws["A3"] = "기본 정보"
            ws["A3"].font = openpyxl.styles.Font(bold=True)
            
            mode_name = "일반 양정" if result["mode"] == "standard" else "압송 관로"
            ws["A4"] = "계산 모드"
            ws["B4"] = mode_name
            
            electrical_str = "삼상 380V" if result["electrical_type"] == "3phase" else "단상 220V"
            ws["A5"] = "전기 방식"
            ws["B5"] = electrical_str
            
            ws["A6"] = "펌프 대수"
            ws["B6"] = result.get("num_pumps", 1)
            
            # 계산 결과
            ws["A8"] = "계산 결과"
            ws["A8"].font = openpyxl.styles.Font(bold=True)
            
            data_rows = [
                ("전양정", f"{result['total_head']:.2f} m"),
                ("펌프당 유량", f"{result['flow_per_pump']:.2f} ㎥/hr"),
                ("펌프 동력", f"{result['pump_power']:.2f} kW"),
                ("펌프 효율", f"{result['efficiency']*100:.0f}%"),
                ("모터 용량", f"{result['motor_power']:.2f} kW"),
                ("모터 여유율", f"{result['motor_safety_factor']}"),
                ("전류", f"{result['current']:.2f} A"),
            ]
            
            for i, (label, value) in enumerate(data_rows, start=9):
                ws[f"A{i}"] = label
                ws[f"B{i}"] = value
            
            # 압송 모드 추가 정보
            if result["mode"] == "pressure" and "friction_loss" in result:
                ws["A17"] = "압송 관로 상세"
                ws["A17"].font = openpyxl.styles.Font(bold=True)
                
                pressure_data = [
                    ("마찰 손실", f"{result['friction_loss']:.2f} m"),
                    ("잔압 환산", f"{result['residual_head']:.2f} m"),
                    ("유속", f"{result['velocity']:.2f} m/s"),
                ]
                
                for i, (label, value) in enumerate(pressure_data, start=18):
                    ws[f"A{i}"] = label
                    ws[f"B{i}"] = value
            
            # 열 너비 조정
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            
            # 파일 저장
            wb.save(filepath)
            return True
            
        except Exception as e:
            raise e
    
    def save_to_text(self, result, filepath):
        """텍스트 파일로 저장"""
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                # 제목
                f.write("=" * 50 + " ")
                f.write("하수도 펌프 용량 계산 결과
")
                f.write("=" * 50 + "  ")
                
                # 기본 정보
                f.write("[기본 정보]
")
                mode_name = "일반 양정" if result["mode"] == "standard" else "압송 관로"
                f.write(f"계산 모드: {mode_name}\n")
                
                electrical_str = "삼상 380V" if result["electrical_type"] == "3phase" else "단상 220V"
                f.write(f"전기 방식: {electrical_str}\n")
                f.write(f"펌프 대수: {result.get('num_pumps', 1)}대 \n")
                
                # 계산 결과
                f.write("[계산 결과]
")
                f.write(f"전양정: {result['total_head']:.2f} m\n")
                f.write(f"펌프당 유량: {result['flow_per_pump']:.2f} ㎥/hr\n")
                f.write(f"펌프 동력: {result['pump_power']:.2f} kW\n")
                f.write(f"펌프 효율: {result['efficiency']*100:.0f}%\n")
                f.write(f"모터 용량: {result['motor_power']:.2f} kW\n")
                f.write(f"모터 여유율: {result['motor_safety_factor']}\n")
                f.write(f"전류: {result['current']:.2f} A \n")
                
                # 압송 모드 추가 정보
                if result["mode"] == "pressure" and "friction_loss" in result:
                    f.write("[압송 관로 상세]
")
                    f.write(f"마찰 손실: {result['friction_loss']:.2f} m\n")
                    f.write(f"잔압 환산: {result['residual_head']:.2f} m\n")
                    f.write(f"유속: {result['velocity']:.2f} m/s \n")
                
                # 저장 정보
                f.write("-" * 50 + "
")
                f.write(f"저장 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"파일 형식: {'Excel' if self.file_format_var.get() == 'excel' else 'Text'}\n")
            
            return True
            
        except Exception as e:
            raise e
    

def create_left_panel(self, parent):
        """왼쪽 패널 (입력 영역) 생성"""
        # 모드 선택
        self.create_mode_selection(parent)

        # 기본 정보 입력
        self.create_basic_inputs(parent)

        # 전기 사양 선택
        self.create_electrical_selection(parent)
        
        # 파일 저장 설정
        self.create_file_save_section(parent)
        

        # 버튼 영역
        self.create_buttons(parent)

    
    def create_file_save_section(self, parent):
        """파일 저장 설정 영역 생성"""
        save_frame = ttk.Frame(parent)
        save_frame.pack(fill=tk.X, padx=10, pady=15)
        
        # 제목 라벨
        title_label = tk.Label(
            save_frame,
            text="파일 저장 설정",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(anchor=tk.W, padx=5, pady=(0, 10))
        
        # 내용 프레임
        content_frame = ttk.Frame(save_frame)
        content_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 저장 경로 선택
        tk.Label(content_frame, text="저장 위치:", font=("Helvetica", 13)).grid(
            row=0, column=0, padx=8, pady=8, sticky=tk.W
        )
        
        # 경로 표시 Entry
        path_entry = tk.Entry(content_frame, textvariable=self.save_path_var, width=40, font=("Helvetica", 12))
        path_entry.grid(row=0, column=1, padx=8, pady=8)
        
        # 폴더 선택 버튼
        browse_button = tk.Button(
            content_frame,
            text="폴더 선택",
            command=self.browse_save_folder,
            bg="#95a5a6",
            fg="white",
            font=("Helvetica", 12),
            padx=15,
            pady=6
        )
        browse_button.grid(row=0, column=2, padx=8, pady=8)
        
        # 파일 형식 선택
        tk.Label(content_frame, text="파일 형식:", font=("Helvetica", 13)).grid(
            row=1, column=0, padx=8, pady=8, sticky=tk.W
        )
        
        # 라디오 버튼 프레임
        format_frame = ttk.Frame(content_frame)
        format_frame.grid(row=1, column=1, columnspan=2, padx=8, pady=8, sticky=tk.W)
        
        # Excel 형식
        excel_radio = tk.Radiobutton(
            format_frame,
            text="Excel 파일 (.xlsx)",
            variable=self.file_format_var,
            value="excel",
            font=("Helvetica", 12)
        )
        excel_radio.pack(side=tk.LEFT, padx=15)
        
        # 텍스트 형식
        text_radio = tk.Radiobutton(
            format_frame,
            text="텍스트 파일 (.txt)",
            variable=self.file_format_var,
            value="text",
            font=("Helvetica", 12)
        )
        text_radio.pack(side=tk.LEFT, padx=15)
        
        return save_frame
    
    def browse_save_folder(self):
        """저장 폴더 선택 다이얼로그"""
        from tkinter import filedialog
        folder_path = filedialog.askdirectory(
            title="결과 파일 저장 폴더 선택",
            initialdir=self.save_path_var.get()
        )
        if folder_path:
            self.save_path_var.set(folder_path)
    
    def save_calculation_result(self, result):
        """계산 결과를 파일로 저장"""
        try:
            import os
            import datetime
            
            # 파일명 생성
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            mode_name = "일반양정" if result["mode"] == "standard" else "압송관로"
            
            if self.file_format_var.get() == "excel":
                filename = f"펌프계산_{mode_name}_{timestamp}.xlsx"
                filepath = os.path.join(self.save_path_var.get(), filename)
                self.save_to_excel(result, filepath)
            else:
                filename = f"펌프계산_{mode_name}_{timestamp}.txt"
                filepath = os.path.join(self.save_path_var.get(), filename)
                self.save_to_text(result, filepath)
            
            # 저장 완료 메시지
            from tkinter import messagebox
            messagebox.showinfo("저장 완료", f"계산 결과가 저장되었습니다:\n{filepath}")
            
            return True
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("저장 오류", f"파일 저장 중 오류가 발생했습니다:\n{str(e)}")
            return False
    
    def save_to_excel(self, result, filepath):
        """Excel 파일로 저장"""
        try:
            # openpyxl이 설치되어 있는지 확인
            try:
                import openpyxl
                from openpyxl import Workbook
            except ImportError:
                # openpyxl이 없으면 설치하도록 안내
                from tkinter import messagebox
                messagebox.showwarning(
                    "라이브러리 필요",
                    "Excel 저장을 위해 openpyxl 라이브러리가 필요합니다.\n\n설치 명령: pip install openpyxl"
                )
                return False
            
            # 새 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "펌프 계산 결과"
            
            # 제목
            ws["A1"] = "하수도 펌프 용량 계산 결과"
            ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
            
            # 기본 정보
            ws["A3"] = "기본 정보"
            ws["A3"].font = openpyxl.styles.Font(bold=True)
            
            mode_name = "일반 양정" if result["mode"] == "standard" else "압송 관로"
            ws["A4"] = "계산 모드"
            ws["B4"] = mode_name
            
            electrical_str = "삼상 380V" if result["electrical_type"] == "3phase" else "단상 220V"
            ws["A5"] = "전기 방식"
            ws["B5"] = electrical_str
            
            ws["A6"] = "펌프 대수"
            ws["B6"] = result.get("num_pumps", 1)
            
            # 계산 결과
            ws["A8"] = "계산 결과"
            ws["A8"].font = openpyxl.styles.Font(bold=True)
            
            data_rows = [
                ("전양정", f"{result['total_head']:.2f} m"),
                ("펌프당 유량", f"{result['flow_per_pump']:.2f} ㎥/hr"),
                ("펌프 동력", f"{result['pump_power']:.2f} kW"),
                ("펌프 효율", f"{result['efficiency']*100:.0f}%"),
                ("모터 용량", f"{result['motor_power']:.2f} kW"),
                ("모터 여유율", f"{result['motor_safety_factor']}"),
                ("전류", f"{result['current']:.2f} A"),
            ]
            
            for i, (label, value) in enumerate(data_rows, start=9):
                ws[f"A{i}"] = label
                ws[f"B{i}"] = value
            
            # 압송 모드 추가 정보
            if result["mode"] == "pressure" and "friction_loss" in result:
                ws["A17"] = "압송 관로 상세"
                ws["A17"].font = openpyxl.styles.Font(bold=True)
                
                pressure_data = [
                    ("마찰 손실", f"{result['friction_loss']:.2f} m"),
                    ("잔압 환산", f"{result['residual_head']:.2f} m"),
                    ("유속", f"{result['velocity']:.2f} m/s"),
                ]
                
                for i, (label, value) in enumerate(pressure_data, start=18):
                    ws[f"A{i}"] = label
                    ws[f"B{i}"] = value
            
            # 열 너비 조정
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            
            # 파일 저장
            wb.save(filepath)
            return True
            
        except Exception as e:
            raise e
    
    def save_to_text(self, result, filepath):
        """텍스트 파일로 저장"""
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                # 제목
                f.write("=" * 50 + " ")
                f.write("하수도 펌프 용량 계산 결과
")
                f.write("=" * 50 + "  ")
                
                # 기본 정보
                f.write("[기본 정보]
")
                mode_name = "일반 양정" if result["mode"] == "standard" else "압송 관로"
                f.write(f"계산 모드: {mode_name}\n")
                
                electrical_str = "삼상 380V" if result["electrical_type"] == "3phase" else "단상 220V"
                f.write(f"전기 방식: {electrical_str}\n")
                f.write(f"펌프 대수: {result.get('num_pumps', 1)}대 \n")
                
                # 계산 결과
                f.write("[계산 결과]
")
                f.write(f"전양정: {result['total_head']:.2f} m\n")
                f.write(f"펌프당 유량: {result['flow_per_pump']:.2f} ㎥/hr\n")
                f.write(f"펌프 동력: {result['pump_power']:.2f} kW\n")
                f.write(f"펌프 효율: {result['efficiency']*100:.0f}%\n")
                f.write(f"모터 용량: {result['motor_power']:.2f} kW\n")
                f.write(f"모터 여유율: {result['motor_safety_factor']}\n")
                f.write(f"전류: {result['current']:.2f} A \n")
                
                # 압송 모드 추가 정보
                if result["mode"] == "pressure" and "friction_loss" in result:
                    f.write("[압송 관로 상세]
")
                    f.write(f"마찰 손실: {result['friction_loss']:.2f} m\n")
                    f.write(f"잔압 환산: {result['residual_head']:.2f} m\n")
                    f.write(f"유속: {result['velocity']:.2f} m/s \n")
                
                # 저장 정보
                f.write("-" * 50 + "
")
                f.write(f"저장 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"파일 형식: {'Excel' if self.file_format_var.get() == 'excel' else 'Text'}\n")
            
            return True
            
        except Exception as e:
            raise e
    

def create_mode_selection(self, parent):
        """계산 모드 선택 영역 생성"""
        mode_frame = ttk.Frame(parent)
        mode_frame.pack(fill=tk.X, padx=10, pady=8)

        # 제목 라벨 추가
        title_label = tk.Label(
            mode_frame,
            text="계산 모드 선택",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # 내용 프레임
        content_frame = ttk.Frame(mode_frame)
        content_frame.pack(fill=tk.X, padx=5, pady=5)

        # 라디오 버튼 프레임
        radio_frame = ttk.Frame(content_frame)
        radio_frame.pack()

        # 일반 양정 모드
        standard_radio = tk.Radiobutton(
            radio_frame,
            text="일반 양정 모드",
            variable=self.mode_var,
            value='standard',
            font=("Helvetica", 13)
        )
        standard_radio.grid(row=0, column=0, padx=25, pady=8, sticky=tk.W)

        # 압송 관로 모드
        pressure_radio = tk.Radiobutton(
            radio_frame,
            text="압송 관로 모드",
            variable=self.mode_var,
            value='pressure',
            font=("Helvetica", 13)
        )
        pressure_radio.grid(row=0, column=1, padx=25, pady=8, sticky=tk.W)

    def create_basic_inputs(self, parent):
        """기본 정보 입력 영역 생성"""
        basic_frame = ttk.Frame(parent)
        basic_frame.pack(fill=tk.X, padx=10, pady=8)

        # 제목 라벨 추가
        title_label = tk.Label(
            basic_frame,
            text="기본 정보 입력",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # 내용 프레임
        content_frame = ttk.Frame(basic_frame)
        content_frame.pack(fill=tk.X, padx=5, pady=5)

        # 설계 유량
        tk.Label(content_frame, text="설계 유량 (㎥/hr):", font=("Helvetica", 13)).grid(
            row=0, column=0, padx=8, pady=8, sticky=tk.W
        )
        flow_entry = tk.Entry(content_frame, textvariable=self.flow_var, width=15, font=("Helvetica", 14))
        flow_entry.grid(row=0, column=1, padx=8, pady=8)
        tk.Label(content_frame, text="예: 100", font=("Helvetica", 12)).grid(
            row=0, column=2, padx=8, pady=8, sticky=tk.W
        )

        # 실양정
        tk.Label(content_frame, text="실양정 (m):", font=("Helvetica", 13)).grid(
            row=1, column=0, padx=8, pady=8, sticky=tk.W
        )
        head_entry = tk.Entry(content_frame, textvariable=self.head_var, width=15, font=("Helvetica", 14))
        head_entry.grid(row=1, column=1, padx=8, pady=8)
        tk.Label(content_frame, text="예: 20", font=("Helvetica", 12)).grid(
            row=1, column=2, padx=8, pady=8, sticky=tk.W
        )

        # 펌프 대수
        tk.Label(content_frame, text="펌프 대수:", font=("Helvetica", 13)).grid(
            row=2, column=0, padx=8, pady=8, sticky=tk.W
        )
        pumps_spinbox = tk.Spinbox(
            content_frame,
            from_=1,
            to=5,
            textvariable=self.pumps_var,
            width=13,
            font=("Helvetica", 14)
        )
        pumps_spinbox.grid(row=2, column=1, padx=8, pady=8)

        # 일반 양정 모드: 관로 손실
        tk.Label(content_frame, text="관로 손실 (m):", font=("Helvetica", 13)).grid(
            row=3, column=0, padx=5, pady=5, sticky=tk.W
        )
        pipe_loss_entry = tk.Entry(content_frame, textvariable=self.pipe_loss_var, font=("Helvetica", 14), width=15)
        pipe_loss_entry.grid(row=3, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="일반 양정 모드용", font=("Helvetica", 13)).grid(
            row=3, column=2, padx=5, pady=5, sticky=tk.W
        )

        # 압송 관로 모드: 관 길이
        tk.Label(content_frame, text="압송 거리 (m):", font=("Helvetica", 13)).grid(
            row=4, column=0, padx=5, pady=5, sticky=tk.W
        )
        pipe_length_entry = tk.Entry(content_frame, textvariable=self.pipe_length_var, font=("Helvetica", 14), width=15)
        pipe_length_entry.grid(row=4, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="압송 모드용", font=("Helvetica", 13)).grid(
            row=4, column=2, padx=5, pady=5, sticky=tk.W
        )

        # 압송 관로 모드: 관경
        tk.Label(content_frame, text="관경 (mm):", font=("Helvetica", 13)).grid(
            row=5, column=0, padx=5, pady=5, sticky=tk.W
        )
        pipe_diameter_entry = tk.Entry(content_frame, textvariable=self.pipe_diameter_var, font=("Helvetica", 14), width=15)
        pipe_diameter_entry.grid(row=5, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="예: 150", font=("Helvetica", 13)).grid(
            row=5, column=2, padx=5, pady=5, sticky=tk.W
        )

        # 압송 관로 모드: 관 재질
        tk.Label(content_frame, text="관 재질:", font=("Helvetica", 13)).grid(
            row=6, column=0, padx=8, pady=8, sticky=tk.W
        )
        material_combo = ttk.Combobox(
            content_frame,
            textvariable=self.pipe_material_var,
            values=['PVC', 'PE', '강관', 'HDPE', '주철관'],
            width=12,
            state='readonly'
        )
        material_combo.grid(row=6, column=1, padx=8, pady=8)

        # 옵션: 여유 수두
        tk.Label(content_frame, text="여유 수두 (m):", font=("Helvetica", 13)).grid(
            row=7, column=0, padx=5, pady=5, sticky=tk.W
        )
        safety_entry = tk.Entry(content_frame, textvariable=self.safety_var, font=("Helvetica", 14), width=15)
        safety_entry.grid(row=7, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="기본값: 1.5", font=("Helvetica", 13)).grid(
            row=7, column=2, padx=5, pady=5, sticky=tk.W
        )

        # 옵션: 펌프 효율
        tk.Label(content_frame, text="펌프 효율:", font=("Helvetica", 13)).grid(
            row=8, column=0, padx=5, pady=5, sticky=tk.W
        )
        efficiency_entry = tk.Entry(content_frame, textvariable=self.efficiency_var, font=("Helvetica", 14), width=15)
        efficiency_entry.grid(row=8, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="기본값: 0.70", font=("Helvetica", 13)).grid(
            row=8, column=2, padx=5, pady=5, sticky=tk.W
        )

        # 옵션: 모터 여유율
        tk.Label(content_frame, text="모터 여유율:", font=("Helvetica", 13)).grid(
            row=9, column=0, padx=5, pady=5, sticky=tk.W
        )
        motor_safety_entry = tk.Entry(content_frame, textvariable=self.motor_safety_var, font=("Helvetica", 14), width=15)
        motor_safety_entry.grid(row=9, column=1, padx=5, pady=5)
        tk.Label(content_frame, text="기본값: 1.15", font=("Helvetica", 13)).grid(
            row=9, column=2, padx=5, pady=5, sticky=tk.W
        )

    def create_electrical_selection(self, parent):
        """전기 사양 선택 영역 생성"""
        electrical_frame = ttk.Frame(parent)
        electrical_frame.pack(fill=tk.X, padx=10, pady=8)

        # 제목 라벨 추가
        title_label = tk.Label(
            electrical_frame,
            text="전기 사양 선택",
            font=("Helvetica", 14, "bold")
        )
        title_label.pack(anchor=tk.W, padx=5, pady=(0, 10))

        # 내용 프레임
        content_frame = ttk.Frame(electrical_frame)
        content_frame.pack(fill=tk.X, padx=5, pady=5)

        # 라디오 버튼 프레임
        radio_frame = ttk.Frame(content_frame)
        radio_frame.pack()

        # 삼상 전기
        phase3_radio = tk.Radiobutton(
            radio_frame,
            text="삼상 380V",
            variable=self.electrical_var,
            value='3phase',
            font=("Helvetica", 13)
        )
        phase3_radio.grid(row=0, column=0, padx=25, pady=8, sticky=tk.W)
        
        # 단상 전기
        phase1_radio = tk.Radiobutton(
            radio_frame,
            text="단상 220V",
            variable=self.electrical_var,
            value='1phase',
            font=("Helvetica", 13)
        )
        phase1_radio.grid(row=0, column=1, padx=25, pady=8, sticky=tk.W)

    def create_buttons(self, parent):
        """버튼 영역 생성"""
        button_frame = ttk.Frame(parent, padding="10")
        button_frame.pack(fill=tk.X)

        # 계산하기 버튼 (강조)
        self.calc_button = tk.Button(
            button_frame,
            text="계산하기",
            command=self.calculate,
            bg='#3498db',
            fg='white',
            font=('Helvetica', 14, 'bold'),
            padx=25,
            pady=12,
            relief=tk.RAISED,
            borderwidth=2
        )
        self.calc_button.pack(side=tk.LEFT, padx=15)

        # 새 계산 버튼
        reset_button = tk.Button(
            button_frame,
            text="새 계산",
            command=self.reset_inputs,
            bg='#95a5a6',
            fg='white',
            font=('Helvetica', 14),
            padx=25,
            pady=12,
            relief=tk.RAISED,
            borderwidth=2
        )
        reset_button.pack(side=tk.LEFT, padx=15)

        # 종료 버튼
        exit_button = tk.Button(
            button_frame,
            text="종료",
            command=self.root.quit,
            bg='#e74c3c',
            fg='white',
            font=('Helvetica', 14),
            padx=25,
            pady=12,
            relief=tk.RAISED,
            borderwidth=2
        )
        exit_button.pack(side=tk.RIGHT, padx=15)

    def on_mode_changed(self, *args):
        """모드 변경 시 호출되는 함수"""
        mode = self.mode_var.get()
        # TODO: 모드에 따라 입력 필드 활성화/비활성화
        # Phase 2에서 구현 예정
        pass

        def calculate_and_save(self):
        """계산 실행 및 결과 파일 저장"""
        try:
            # 입력값 검증
            flow = self.flow_var.get().strip()
            head = self.head_var.get().strip()
            
            if not flow or not head:
                from tkinter import messagebox
                messagebox.showerror("입력 오류", "설계 유량과 실양정을 입력해주세요.")
                return
            
            # 계산 실행 (기존 calculate 로직)
            # ... (기존 calculate 함수 내용 유지)
            
            # 임시: 계산 성공 메시지
            from tkinter import messagebox
            messagebox.showinfo("계산 완료", "계산이 완료되었습니다.
결과 파일이 저장됩니다.")
            
            # 파일 저장 호출 (임시)
            # 실제로는 계산 결과를 얻어서 save_calculation_result 호출
            print("계산 및 저장 기능이 준비되었습니다.")
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("계산 오류", f"오류 발생: {str(e)}")

    def show_error(self, message):
        """에러 메시지 표시"""
        from tkinter import messagebox
        messagebox.showerror("오류", message)

    def insert_result_line(self, label, value, highlight=False):
        """결과 라인 삽입 헬퍼 함수"""
        self.result_text.insert(tk.END, f" {label}: ", 'label')
        if highlight:
            self.result_text.insert(tk.END, f"{value}\n", 'highlight')
        else:
            self.result_text.insert(tk.END, f"{value}\n", 'value')

    def reset_inputs(self):
        """입력 필드 초기화"""
        self.flow_var.set('')
        self.head_var.set('')
        self.pumps_var.set(1)
        self.pipe_loss_var.set('0')
        self.pipe_length_var.set('')
        self.pipe_diameter_var.set('')
        self.pipe_material_var.set('PVC')
        self.safety_var.set('1.5')
        self.efficiency_var.set('0.70')
        self.motor_safety_var.set('1.15')
        self.mode_var.set('standard')
        self.electrical_var.set('3phase')

        # 결과 영역 초기화
        welcome_msg = """


 👋 환영합니다!

 왼쪽 입력 폼을 채우고
 '계산하기' 버튼을 눌러주세요.

 계산 결과가 여기에 표시됩니다.


"""
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete('1.0', tk.END)
        self.result_text.insert('1.0', welcome_msg, 'section')
        self.result_text.config(state=tk.DISABLED)


def main():
    """메인 함수"""
    root = tk.Tk()
    app = PumpCalculatorGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()